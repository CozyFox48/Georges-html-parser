import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from copy import copy

excel_file = 'extracted.xlsx'

wb = openpyxl.load_workbook(excel_file)
summary = wb['Summary']
tabular = wb['Tabular']

def auto_convert_to_number(string):
    try:
        number = int(string)
        return number
    except ValueError:
        try:
            number = float(string)
            return number
        except ValueError:
            return string
    except TypeError:
    	return string

def append_summary(row_number, row_values):

	font = Font(name='Times New Roman', size=12, bold=False)
	alignment = Alignment(vertical='center', wrapText=True)

	for index, value in enumerate(row_values):
		cell = summary.cell(row=row_number, column=index+1)
		cell.value = auto_convert_to_number(value)
		cell.font = font
		cell.alignment = alignment

def add_sheet(sheet_name, chemical_composition, experimental_data):

	new_sheet = wb.create_sheet(sheet_name)

	for row in tabular.iter_rows(values_only=True):
		new_sheet.append(row)

	new_sheet.column_dimensions['A'].width = tabular.column_dimensions['A'].width
	new_sheet.column_dimensions['B'].width = tabular.column_dimensions['B'].width
	new_sheet.column_dimensions['C'].width = tabular.column_dimensions['C'].width
	new_sheet.column_dimensions['D'].width = tabular.column_dimensions['D'].width
	new_sheet.column_dimensions['E'].width = tabular.column_dimensions['E'].width

	for i in range(1, 36+1):
		for j in range(1, 5+1):

			cell_src = tabular.cell(row=i, column=j)

			cell = new_sheet.cell(row=i, column=j)
			cell.value = auto_convert_to_number(cell_src.value)
			cell.fill = copy(cell_src.fill)
			cell.font = copy(cell_src.font)
			cell.border = copy(cell_src.border)
			cell.alignment = copy(cell_src.alignment)

	font = Font(name='Times New Roman', size=12, bold=False)
	alignment = Alignment(horizontal='right')
	border = Border(
		left=Side(border_style="thin", color="000000"),
		right=Side(border_style="thin", color="000000"),
		top=Side(border_style="thin", color="000000"),
		bottom=Side(border_style="thin", color="000000")
	)

	for _ in range(len(chemical_composition)):

		for index, value in enumerate(chemical_composition[_]):
			cell = new_sheet.cell(row=_+5, column=index+1)
			cell.value = auto_convert_to_number(value)
			cell.font = font

	for _ in range(len(experimental_data)):

		for index, value in enumerate(experimental_data[_]):
			cell = new_sheet.cell(row=_+23, column=index+1)
			cell.value = auto_convert_to_number(value)
			cell.font = font
			cell.border = border
			cell.alignment = alignment

def remove_tabular():

	wb.remove(tabular)
	wb.save(excel_file)